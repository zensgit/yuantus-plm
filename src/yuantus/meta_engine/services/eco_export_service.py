from __future__ import annotations

import csv
import io
import json
from typing import Any, Dict, Iterable, List, Optional


class ImpactExportSection:
    def __init__(self, name: str, columns: List[str], rows: List[List[Any]]):
        self.name = name
        self.columns = columns
        self.rows = rows


class EcoImpactExportService:
    def __init__(self, impact: Dict[str, Any]):
        self.impact = impact

    def build_sections(self) -> List[ImpactExportSection]:
        sections: List[ImpactExportSection] = []

        bom_diff_meta = self.impact.get("bom_diff") or {}

        overview_rows = [
            ["eco_id", self.impact.get("eco_id")],
            ["changed_product_id", self.impact.get("changed_product_id")],
            ["impact_count", self.impact.get("impact_count")],
            ["impact_level", self.impact.get("impact_level")],
            ["impact_score", self.impact.get("impact_score")],
            ["impact_scope", self.impact.get("impact_scope")],
        ]
        if bom_diff_meta.get("compare_mode") is not None:
            overview_rows.append(["bom_compare_mode", bom_diff_meta.get("compare_mode")])
        if bom_diff_meta.get("line_key") is not None:
            overview_rows.append(["bom_line_key", bom_diff_meta.get("line_key")])
        sections.append(ImpactExportSection("Overview", ["key", "value"], overview_rows))

        impact_summary = self.impact.get("impact_summary") or {}
        summary_rows = [[k, impact_summary.get(k)] for k in sorted(impact_summary.keys())]
        sections.append(
            ImpactExportSection("Impact Summary", ["metric", "value"], summary_rows)
        )

        files_summary = self.impact.get("files_summary") or {}
        if files_summary:
            file_rows = [
                [k, files_summary.get(k)] for k in sorted(files_summary.keys())
            ]
            sections.append(
                ImpactExportSection("Files Summary", ["metric", "value"], file_rows)
            )

        impacted_rows: List[List[Any]] = []
        for entry in self.impact.get("impacted_assemblies") or []:
            parent = entry.get("parent") or {}
            rel = entry.get("relationship") or {}
            impacted_rows.append(
                [
                    parent.get("id"),
                    self._item_field(parent, "item_number"),
                    self._item_field(parent, "name"),
                    entry.get("level"),
                    rel.get("id"),
                    rel.get("item_type_id"),
                ]
            )
        sections.append(
            ImpactExportSection(
                "Impacted Assemblies",
                [
                    "parent_id",
                    "parent_item_number",
                    "parent_name",
                    "level",
                    "relationship_id",
                    "relationship_type",
                ],
                impacted_rows,
            )
        )

        bom_diff = self.impact.get("bom_diff") or {}
        sections.extend(self._bom_diff_sections(bom_diff))

        version_diff = self.impact.get("version_diff") or {}
        if version_diff:
            v_rows = []
            for field, values in sorted((version_diff.get("diffs") or {}).items()):
                v_rows.append(
                    [
                        version_diff.get("version_a"),
                        version_diff.get("version_b"),
                        field,
                        (values or {}).get("a"),
                        (values or {}).get("b"),
                    ]
                )
            sections.append(
                ImpactExportSection(
                    "Version Diff",
                    ["version_a", "version_b", "field", "left", "right"],
                    v_rows,
                )
            )

        version_files_diff = self.impact.get("version_files_diff") or {}
        if version_files_diff:
            vf_rows = []
            for change_type in ("added", "removed", "modified"):
                for entry in version_files_diff.get(change_type) or []:
                    vf_rows.append(
                        [
                            change_type,
                            entry.get("file_id"),
                            entry.get("file_role"),
                            entry.get("filename"),
                            entry.get("checksum_a"),
                            entry.get("checksum_b"),
                        ]
                    )
            sections.append(
                ImpactExportSection(
                    "Version Files Diff",
                    [
                        "change_type",
                        "file_id",
                        "file_role",
                        "filename",
                        "checksum_a",
                        "checksum_b",
                    ],
                    vf_rows,
                )
            )

        files = self.impact.get("files") or {}
        sections.extend(self._file_sections("Item Files", files.get("item_files")))
        sections.extend(
            self._file_sections(
                "Source Version Files", files.get("source_version_files")
            )
        )
        sections.extend(
            self._file_sections(
                "Target Version Files", files.get("target_version_files")
            )
        )

        return sections

    def to_csv(self) -> str:
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for section in self.build_sections():
            output.write(f"# {section.name}\n")
            writer.writerow(section.columns)
            for row in section.rows:
                writer.writerow([self._format_cell(v) for v in row])
            output.write("\n")
        return output.getvalue()

    def to_xlsx(self) -> bytes:
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - guarded by dependency
            raise RuntimeError("openpyxl is required for xlsx export") from exc

        wb = Workbook()
        wb.remove(wb.active)
        for section in self.build_sections():
            title = self._sanitize_sheet_name(section.name)
            ws = wb.create_sheet(title=title)
            ws.append(section.columns)
            for row in section.rows:
                ws.append([self._format_cell(v) for v in row])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def to_pdf(self) -> bytes:
        try:
            from fpdf import FPDF
        except ImportError as exc:  # pragma: no cover - guarded by dependency
            raise RuntimeError("fpdf2 is required for pdf export") from exc

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=12)
        pdf.add_page()
        pdf.set_font("Helvetica", size=10)

        max_width = pdf.w - pdf.l_margin - pdf.r_margin
        for section in self.build_sections():
            pdf.set_font("Helvetica", style="B", size=11)
            pdf.cell(0, 8, section.name, ln=1)
            pdf.set_font("Helvetica", size=9)
            header_line = self._join_line(section.columns)
            for line in self._wrap_pdf_lines(pdf, header_line, max_width):
                pdf.cell(max_width, 5, line, ln=1)
            for row in section.rows:
                for line in self._wrap_pdf_lines(
                    pdf, self._join_line(row), max_width
                ):
                    pdf.cell(max_width, 5, line, ln=1)
            pdf.ln(2)

        data = pdf.output(dest="S")
        if isinstance(data, bytearray):
            return bytes(data)
        if isinstance(data, str):
            return data.encode("latin-1")
        return data

    def _bom_diff_sections(self, bom_diff: Dict[str, Any]) -> List[ImpactExportSection]:
        sections: List[ImpactExportSection] = []
        if not bom_diff:
            return sections

        def entry_rows(entries: Iterable[Dict[str, Any]]) -> List[List[Any]]:
            rows = []
            for entry in entries:
                parent = entry.get("parent") or {}
                child = entry.get("child") or {}
                rows.append(
                    [
                        entry.get("parent_id"),
                        self._item_field(parent, "item_number"),
                        entry.get("child_id"),
                        self._item_field(child, "item_number"),
                        entry.get("relationship_id"),
                        entry.get("line_key"),
                        entry.get("level"),
                        self._safe_json(entry.get("properties")),
                    ]
                )
            return rows

        added = entry_rows(bom_diff.get("added") or [])
        removed = entry_rows(bom_diff.get("removed") or [])
        if added:
            sections.append(
                ImpactExportSection(
                    "BOM Diff Added",
                    [
                        "parent_id",
                        "parent_item_number",
                        "child_id",
                        "child_item_number",
                        "relationship_id",
                        "line_key",
                        "level",
                        "properties",
                    ],
                    added,
                )
            )
        if removed:
            sections.append(
                ImpactExportSection(
                    "BOM Diff Removed",
                    [
                        "parent_id",
                        "parent_item_number",
                        "child_id",
                        "child_item_number",
                        "relationship_id",
                        "line_key",
                        "level",
                        "properties",
                    ],
                    removed,
                )
            )

        changed_rows: List[List[Any]] = []
        for entry in bom_diff.get("changed") or []:
            changes = entry.get("changes") or []
            for change in changes:
                changed_rows.append(
                    [
                        entry.get("parent_id"),
                        entry.get("child_id"),
                        entry.get("relationship_id"),
                        entry.get("line_key"),
                        entry.get("level"),
                        change.get("field"),
                        change.get("left"),
                        change.get("right"),
                        change.get("severity"),
                    ]
                )
        if changed_rows:
            sections.append(
                ImpactExportSection(
                    "BOM Diff Changed",
                    [
                        "parent_id",
                        "child_id",
                        "relationship_id",
                        "line_key",
                        "level",
                        "field",
                        "left",
                        "right",
                        "severity",
                    ],
                    changed_rows,
                )
            )

        return sections

    def _file_sections(
        self, title: str, entries: Optional[Iterable[Dict[str, Any]]]
    ) -> List[ImpactExportSection]:
        if not entries:
            return []
        rows = []
        for entry in entries:
            rows.append(
                [
                    entry.get("file_id"),
                    entry.get("file_role"),
                    entry.get("filename"),
                    entry.get("file_type"),
                    entry.get("file_size"),
                    entry.get("sequence"),
                    entry.get("is_primary"),
                ]
            )
        return [
            ImpactExportSection(
                title,
                [
                    "file_id",
                    "file_role",
                    "filename",
                    "file_type",
                    "file_size",
                    "sequence",
                    "is_primary",
                ],
                rows,
            )
        ]

    def _item_field(self, item: Dict[str, Any], field: str) -> Any:
        if field in item:
            return item.get(field)
        return (item.get("properties") or {}).get(field)

    def _safe_json(self, value: Any) -> str:
        if value is None:
            return ""
        try:
            return json.dumps(value, ensure_ascii=False)
        except TypeError:
            return str(value)

    def _format_cell(self, value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return self._safe_json(value)
        return value

    def _sanitize_sheet_name(self, name: str) -> str:
        invalid = set('[]:*?/\\')
        cleaned = "".join(c for c in name if c not in invalid)
        return (cleaned or "Sheet")[:31]

    def _join_line(self, values: Iterable[Any]) -> str:
        return " | ".join(str(self._format_cell(v) or "") for v in values)

    def _wrap_pdf_lines(self, pdf, line: str, max_width: float) -> List[str]:
        if not line:
            return [""]
        return self._split_pdf_line(pdf, line, max_width)

    def _split_pdf_line(self, pdf, line: str, max_width: float) -> List[str]:
        words = line.split(" ")
        lines: List[str] = []
        current = ""

        def add_current():
            nonlocal current
            if current:
                lines.append(current)
                current = ""

        for word in words:
            candidate = f"{current} {word}".strip()
            if pdf.get_string_width(candidate) <= max_width:
                current = candidate
                continue

            if current:
                add_current()

            # Word longer than line width: split by characters.
            if pdf.get_string_width(word) > max_width:
                chunk = ""
                for ch in word:
                    next_chunk = f"{chunk}{ch}"
                    if pdf.get_string_width(next_chunk) <= max_width:
                        chunk = next_chunk
                        continue
                    if chunk:
                        lines.append(chunk)
                        chunk = ch
                if chunk:
                    lines.append(chunk)
                current = ""
            else:
                current = word

        add_current()
        return lines
