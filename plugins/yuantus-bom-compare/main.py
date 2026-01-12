from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple, TYPE_CHECKING

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response
from pydantic import BaseModel, Field
if TYPE_CHECKING:  # pragma: no cover
    from sqlalchemy.orm import Session
else:
    Session = Any

from yuantus.exceptions.handlers import PLMException

router = APIRouter(prefix="/plugins/bom-compare", tags=["plugins-bom-compare"])


def _get_db():
    from yuantus.database import get_db

    yield from get_db()


def _get_identity_db():
    from yuantus.security.auth.database import get_identity_db

    yield from get_identity_db()


def _current_user_optional(
    request: Request,
    identity_db=Depends(_get_identity_db),
    db=Depends(_get_db),
):
    from yuantus.api.dependencies.auth import get_current_user_optional

    return get_current_user_optional(request, identity_db=identity_db, db=db)


def _current_user(
    request: Request,
    identity_db=Depends(_get_identity_db),
    db=Depends(_get_db),
):
    from yuantus.api.dependencies.auth import (
        get_current_user,
        get_current_user_optional,
    )

    user = get_current_user_optional(request, identity_db=identity_db, db=db)
    return get_current_user(user=user)


class BomCompareRequest(BaseModel):
    item_id_a: str = Field(..., description="Root item id for BOM A")
    item_id_b: str = Field(..., description="Root item id for BOM B")
    compare_mode: Optional[str] = Field(
        default="only_product",
        description=(
            "only_product|summarized|num_qty|by_position|by_reference"
        ),
    )
    relationship_types: Optional[List[str]] = Field(
        default=None,
        description="Relationship ItemType ids to traverse (default: all)",
    )
    quantity_key: str = Field(default="quantity")
    position_key: str = Field(default="find_num")
    refdes_key: str = Field(default="refdes")
    levels: int = Field(default=-1, description="Depth for BOM expansion")
    include_unchanged: bool = Field(default=False)
    quantity_tolerance: Optional[float] = Field(
        default=None, description="Numeric tolerance for quantity comparison"
    )
    filters: Optional["BomCompareFilters"] = None


class BomCompareFilters(BaseModel):
    include_statuses: Optional[List[str]] = None
    exclude_statuses: Optional[List[str]] = None
    include_child_ids: Optional[List[str]] = None
    exclude_child_ids: Optional[List[str]] = None
    child_id_regex: Optional[str] = None
    name_regex: Optional[str] = None
    relationship_id_regex: Optional[str] = None
    min_delta: Optional[float] = None
    max_delta: Optional[float] = None


class BomCompareDiff(BaseModel):
    key: str
    child_id: str
    name: Optional[str] = None
    status: str
    qty_a: Optional[float] = None
    qty_b: Optional[float] = None
    delta: Optional[float] = None
    position_a: Optional[str] = None
    position_b: Optional[str] = None
    refdes_a: Optional[str] = None
    refdes_b: Optional[str] = None
    relationship_ids_a: List[str] = Field(default_factory=list)
    relationship_ids_b: List[str] = Field(default_factory=list)


class BomCompareResponse(BaseModel):
    summary: Dict[str, int]
    differences: List[BomCompareDiff]
    summary_filtered: Optional[Dict[str, int]] = None


class BomCompareExportRequest(BomCompareRequest):
    format: str = Field(default="csv", description="Export format (csv|xlsx)")
    columns: Optional[List[str]] = Field(
        default=None,
        description=(
            "Optional column list. Defaults to key,status,child_id,name,qty_a,qty_b,"
            "delta,position_a,position_b,refdes_a,refdes_b,relationship_ids_a,relationship_ids_b"
        ),
    )
    exclude_columns: Optional[List[str]] = Field(
        default=None, description="Columns to exclude from export"
    )
    delimiter: str = Field(default=",", description="CSV delimiter (single character)")
    filename: Optional[str] = Field(default=None, description="Download filename override")
    diff_only: bool = Field(
        default=False, description="Export only added/removed/modified rows"
    )
    xlsx_sheet_mode: str = Field(
        default="combined",
        description="XLSX sheet layout (combined|split)",
    )


class BomApplyChange(BaseModel):
    op: str = Field(..., description="add|remove|update")
    relationship_id: Optional[str] = None
    parent_id: Optional[str] = None
    child_id: Optional[str] = None
    relationship_type: Optional[str] = None
    quantity: Optional[float] = None
    properties: Dict[str, Any] = Field(default_factory=dict)


class BomApplyRequest(BaseModel):
    relationship_types: Optional[List[str]] = None
    changes: List[BomApplyChange]
    dry_run: bool = Field(
        default=False, description="Validate and return results without applying changes"
    )
    preview: bool = Field(
        default=False, description="Preview changes without applying AML operations"
    )


class BomApplyResponse(BaseModel):
    ok: bool
    results: List[Dict[str, Any]]
    dry_run: bool = False
    preview: bool = False

_EXPORT_COLUMNS = [
    "key",
    "status",
    "child_id",
    "name",
    "qty_a",
    "qty_b",
    "delta",
    "position_a",
    "position_b",
    "refdes_a",
    "refdes_b",
    "relationship_ids_a",
    "relationship_ids_b",
]


def _normalize_compare_mode(mode: Optional[str]) -> str:
    if not mode:
        return "only_product"
    normalized = mode.strip().lower().replace("-", "_")
    if normalized in {"only_product", "only"}:
        return "only_product"
    if normalized in {"summarized", "summary"}:
        return "summarized"
    if normalized in {"num_qty", "numqty"}:
        return "num_qty"
    if normalized in {"by_position", "by_pos", "position"}:
        return "by_position"
    if normalized in {"by_reference", "by_ref", "reference"}:
        return "by_reference"
    raise ValueError(
        "compare_mode must be one of: only_product, summarized, num_qty, "
        "by_position, by_reference"
    )


def _parse_quantity(value: Any) -> float:
    if value is None:
        return 1.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return 1.0


def _float_equal(a: Optional[float], b: Optional[float], tolerance: float = 1e-9) -> bool:
    if a is None or b is None:
        return a is b
    return abs(a - b) <= tolerance


def _iter_bom_relations(node: Dict[str, Any]) -> Iterable[Tuple[Dict[str, Any], Dict[str, Any]]]:
    for child_entry in node.get("children", []) or []:
        rel = child_entry.get("relationship") or {}
        child = child_entry.get("child") or {}
        yield rel, child
        yield from _iter_bom_relations(child)


def _build_key(
    mode: str,
    child_id: str,
    qty: float,
    position: Optional[str],
    refdes: Optional[str],
) -> Tuple[str, ...]:
    if mode in {"only_product", "summarized"}:
        return (child_id,)
    if mode == "num_qty":
        return (child_id, str(qty))
    if mode == "by_position":
        return (child_id, position or "")
    if mode == "by_reference":
        return (child_id, refdes or "")
    return (child_id,)


def _build_key_label(key: Tuple[str, ...]) -> str:
    return ":".join([part for part in key if part is not None])


def _normalize_status_list(values: Optional[Iterable[str]]) -> Optional[set[str]]:
    if not values:
        return None
    allowed = {"added", "removed", "modified", "unchanged"}
    normalized = {str(v).strip().lower() for v in values if v and str(v).strip()}
    unknown = sorted(normalized - allowed)
    if unknown:
        raise ValueError(f"Unknown status values: {', '.join(unknown)}")
    return normalized


def _compile_regex(pattern: Optional[str], field_name: str) -> Optional[re.Pattern]:
    if not pattern:
        return None
    try:
        return re.compile(pattern)
    except re.error as exc:
        raise ValueError(f"Invalid {field_name}_regex: {exc}") from exc


def _summarize_diffs(diffs: List["BomCompareDiff"]) -> Dict[str, int]:
    summary = {"added": 0, "removed": 0, "modified": 0, "unchanged": 0}
    for diff in diffs:
        summary[diff.status] = summary.get(diff.status, 0) + 1
    return summary


def _filter_diffs(
    diffs: List["BomCompareDiff"], filters: "BomCompareFilters"
) -> List["BomCompareDiff"]:
    include_statuses = _normalize_status_list(filters.include_statuses)
    exclude_statuses = _normalize_status_list(filters.exclude_statuses)
    include_child_ids = set(filters.include_child_ids or [])
    exclude_child_ids = set(filters.exclude_child_ids or [])
    child_id_regex = _compile_regex(filters.child_id_regex, "child_id")
    name_regex = _compile_regex(filters.name_regex, "name")
    rel_regex = _compile_regex(filters.relationship_id_regex, "relationship_id")
    min_delta = filters.min_delta
    max_delta = filters.max_delta
    if min_delta is not None and max_delta is not None and min_delta > max_delta:
        raise ValueError("min_delta cannot be greater than max_delta")

    filtered: List[BomCompareDiff] = []
    for diff in diffs:
        if include_statuses and diff.status not in include_statuses:
            continue
        if exclude_statuses and diff.status in exclude_statuses:
            continue
        if include_child_ids and diff.child_id not in include_child_ids:
            continue
        if exclude_child_ids and diff.child_id in exclude_child_ids:
            continue
        if child_id_regex and not child_id_regex.search(diff.child_id or ""):
            continue
        if name_regex and not name_regex.search(diff.name or ""):
            continue
        if rel_regex:
            rel_ids = (diff.relationship_ids_a or []) + (
                diff.relationship_ids_b or []
            )
            if not any(rel_regex.search(str(rid)) for rid in rel_ids if rid):
                continue
        if min_delta is not None:
            if diff.delta is None or diff.delta < min_delta:
                continue
        if max_delta is not None:
            if diff.delta is None or diff.delta > max_delta:
                continue
        filtered.append(diff)
    return filtered


def _extract_entries(
    bom_tree: Dict[str, Any],
    *,
    mode: str,
    quantity_key: str,
    position_key: str,
    refdes_key: str,
) -> Dict[Tuple[str, ...], Dict[str, Any]]:
    entries: Dict[Tuple[str, ...], Dict[str, Any]] = {}

    for rel, child in _iter_bom_relations(bom_tree):
        props = rel.get("properties") or {}
        qty_value = props.get(quantity_key)
        if qty_value is None:
            qty_value = props.get("qty")
        qty = _parse_quantity(qty_value)
        position = props.get(position_key)
        refdes = props.get(refdes_key)
        child_id = child.get("id") or rel.get("related_id")
        if not child_id:
            continue
        child_name = child.get("name") or child.get("item_number")

        key = _build_key(mode, child_id, qty, position, refdes)
        entry = entries.get(key)
        if not entry:
            entry = {
                "child_id": child_id,
                "name": child_name,
                "qty": 0.0,
                "positions": set(),
                "refdes": set(),
                "relationship_ids": [],
            }
            entries[key] = entry

        entry["relationship_ids"].append(rel.get("id"))
        if position:
            entry["positions"].add(str(position))
        if refdes:
            entry["refdes"].add(str(refdes))

        if mode == "only_product":
            continue
        entry["qty"] += qty

    return entries


def _normalize_export_format(value: Optional[str]) -> str:
    normalized = (value or "csv").strip().lower()
    if normalized == "csv":
        return "csv"
    if normalized == "xlsx":
        return "xlsx"
    raise ValueError("format must be 'csv' or 'xlsx'")


def _normalize_export_columns(
    columns: Optional[List[str]],
    exclude_columns: Optional[List[str]] = None,
) -> List[str]:
    if not columns:
        selected = list(_EXPORT_COLUMNS)
    else:
        selected = [c.strip() for c in columns if c and c.strip()]
        if not selected:
            selected = list(_EXPORT_COLUMNS)
    unknown = [c for c in selected if c not in _EXPORT_COLUMNS]
    if unknown:
        raise ValueError(f"Unknown columns: {', '.join(unknown)}")

    if exclude_columns:
        excludes = [c.strip() for c in exclude_columns if c and c.strip()]
        unknown_excludes = [c for c in excludes if c not in _EXPORT_COLUMNS]
        if unknown_excludes:
            raise ValueError(f"Unknown exclude_columns: {', '.join(unknown_excludes)}")
        exclude_set = set(excludes)
        selected = [c for c in selected if c not in exclude_set]

    return selected


def _normalize_sheet_mode(value: Optional[str]) -> str:
    normalized = (value or "combined").strip().lower()
    if normalized in {"combined", "split"}:
        return normalized
    raise ValueError("xlsx_sheet_mode must be 'combined' or 'split'")


def _build_csv_payload(
    diffs: List[BomCompareDiff],
    *,
    columns: List[str],
    delimiter: str,
) -> str:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=delimiter, lineterminator="\n")
    writer.writerow(columns)
    for diff in diffs:
        row: List[Any] = []
        for column in columns:
            value = getattr(diff, column, "")
            if isinstance(value, list):
                value = "|".join([str(v) for v in value if v is not None])
            elif value is None:
                value = ""
            row.append(value)
        writer.writerow(row)
    return output.getvalue()


def _build_xlsx_payload(
    diffs: List[BomCompareDiff],
    *,
    columns: List[str],
    summary: Dict[str, int],
    sheet_mode: str = "combined",
) -> bytes:
    try:
        import openpyxl
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openpyxl is required for xlsx export") from exc

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_sheet.append(["status", "count"])
    for key in ("added", "removed", "modified", "unchanged"):
        summary_sheet.append([key, summary.get(key, 0)])

    if sheet_mode == "combined":
        diff_sheet = workbook.create_sheet("diffs")
        diff_sheet.append(columns)
        for diff in diffs:
            row: List[Any] = []
            for column in columns:
                value = getattr(diff, column, "")
                if isinstance(value, list):
                    value = "|".join([str(v) for v in value if v is not None])
                elif value is None:
                    value = ""
                row.append(value)
            diff_sheet.append(row)
    else:
        grouped: Dict[str, List[BomCompareDiff]] = {
            "added": [],
            "removed": [],
            "modified": [],
            "unchanged": [],
        }
        for diff in diffs:
            grouped.setdefault(diff.status, []).append(diff)
        for status in ("added", "removed", "modified", "unchanged"):
            sheet = workbook.create_sheet(status)
            sheet.append(columns)
            for diff in grouped.get(status, []):
                row = []
                for column in columns:
                    value = getattr(diff, column, "")
                    if isinstance(value, list):
                        value = "|".join([str(v) for v in value if v is not None])
                    elif value is None:
                        value = ""
                    row.append(value)
                sheet.append(row)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _export_filename(name: Optional[str], extension: str) -> str:
    if name:
        return name
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return f"bom_compare_{timestamp}.{extension}"


def compare_bom_trees(
    bom_a: Dict[str, Any],
    bom_b: Dict[str, Any],
    *,
    mode: str,
    quantity_key: str,
    position_key: str,
    refdes_key: str,
    include_unchanged: bool,
    quantity_tolerance: Optional[float] = None,
    filters: Optional[BomCompareFilters] = None,
) -> BomCompareResponse:
    entries_a = _extract_entries(
        bom_a,
        mode=mode,
        quantity_key=quantity_key,
        position_key=position_key,
        refdes_key=refdes_key,
    )
    entries_b = _extract_entries(
        bom_b,
        mode=mode,
        quantity_key=quantity_key,
        position_key=position_key,
        refdes_key=refdes_key,
    )

    summary = {"added": 0, "removed": 0, "modified": 0, "unchanged": 0}
    diffs: List[BomCompareDiff] = []
    tolerance = quantity_tolerance if quantity_tolerance is not None else 1e-9

    keys = set(entries_a.keys()) | set(entries_b.keys())
    for key in sorted(keys, key=_build_key_label):
        entry_a = entries_a.get(key)
        entry_b = entries_b.get(key)

        if not entry_a:
            status = "added"
        elif not entry_b:
            status = "removed"
        else:
            if mode == "only_product":
                status = "unchanged"
            else:
                status = (
                    "modified"
                    if not _float_equal(
                        entry_a["qty"], entry_b["qty"], tolerance=tolerance
                    )
                    else "unchanged"
                )

        summary[status] += 1
        if status == "unchanged" and not include_unchanged:
            continue

        qty_a = entry_a["qty"] if entry_a else None
        qty_b = entry_b["qty"] if entry_b else None
        delta = qty_b - qty_a if qty_a is not None and qty_b is not None else None

        diffs.append(
            BomCompareDiff(
                key=_build_key_label(key),
                child_id=(entry_a or entry_b)["child_id"],
                name=(entry_a or entry_b)["name"],
                status=status,
                qty_a=qty_a,
                qty_b=qty_b,
                delta=delta,
                position_a=", ".join(sorted(entry_a["positions"])) if entry_a and entry_a["positions"] else None,
                position_b=", ".join(sorted(entry_b["positions"])) if entry_b and entry_b["positions"] else None,
                refdes_a=", ".join(sorted(entry_a["refdes"])) if entry_a and entry_a["refdes"] else None,
                refdes_b=", ".join(sorted(entry_b["refdes"])) if entry_b and entry_b["refdes"] else None,
                relationship_ids_a=[rid for rid in (entry_a or {}).get("relationship_ids", []) if rid],
                relationship_ids_b=[rid for rid in (entry_b or {}).get("relationship_ids", []) if rid],
            )
        )

    summary_filtered = None
    if filters:
        filtered = _filter_diffs(diffs, filters)
        summary_filtered = _summarize_diffs(filtered)
        diffs = filtered

    return BomCompareResponse(
        summary=summary, differences=diffs, summary_filtered=summary_filtered
    )


def _identity_from_user(current_user: Any) -> Tuple[Optional[str], List[str]]:
    if not current_user:
        return None, []
    identity = str(getattr(current_user, "id", "")) or None
    roles = list(getattr(current_user, "roles", []) or [])
    if getattr(current_user, "is_superuser", False) and "superuser" not in roles:
        roles.append("superuser")
    return identity, roles


@router.post("/compare", response_model=BomCompareResponse)
def compare_bom(
    req: BomCompareRequest,
    db: Session = Depends(_get_db),
    current_user: Any = Depends(_current_user_optional),
) -> BomCompareResponse:
    from yuantus.meta_engine.services.bom_service import BOMService

    try:
        mode = _normalize_compare_mode(req.compare_mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    bom_service = BOMService(db)
    try:
        bom_a = bom_service.get_bom_structure(
            req.item_id_a,
            levels=req.levels,
            relationship_types=req.relationship_types,
        )
        bom_b = bom_service.get_bom_structure(
            req.item_id_b,
            levels=req.levels,
            relationship_types=req.relationship_types,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    try:
        return compare_bom_trees(
            bom_a,
            bom_b,
            mode=mode,
            quantity_key=req.quantity_key,
            position_key=req.position_key,
            refdes_key=req.refdes_key,
            include_unchanged=req.include_unchanged,
            quantity_tolerance=req.quantity_tolerance,
            filters=req.filters,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/export")
def export_bom(
    req: BomCompareExportRequest,
    db: Session = Depends(_get_db),
    current_user: Any = Depends(_current_user_optional),
) -> Response:
    from yuantus.meta_engine.services.bom_service import BOMService

    try:
        mode = _normalize_compare_mode(req.compare_mode)
        export_format = _normalize_export_format(req.format)
        columns = _normalize_export_columns(req.columns, req.exclude_columns)
        sheet_mode = _normalize_sheet_mode(req.xlsx_sheet_mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if export_format == "csv":
        if not req.delimiter or len(req.delimiter) != 1:
            raise HTTPException(
                status_code=400, detail="delimiter must be a single character"
            )

    bom_service = BOMService(db)
    try:
        bom_a = bom_service.get_bom_structure(
            req.item_id_a,
            levels=req.levels,
            relationship_types=req.relationship_types,
        )
        bom_b = bom_service.get_bom_structure(
            req.item_id_b,
            levels=req.levels,
            relationship_types=req.relationship_types,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    try:
        result = compare_bom_trees(
            bom_a,
            bom_b,
            mode=mode,
            quantity_key=req.quantity_key,
            position_key=req.position_key,
            refdes_key=req.refdes_key,
            include_unchanged=req.include_unchanged,
            quantity_tolerance=req.quantity_tolerance,
            filters=req.filters,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    diffs = result.differences
    if req.diff_only:
        diffs = [diff for diff in diffs if diff.status != "unchanged"]
    export_summary = _summarize_diffs(diffs)

    if export_format == "csv":
        payload = _build_csv_payload(
            diffs, columns=columns, delimiter=req.delimiter
        )
        media_type = "text/csv; charset=utf-8"
        extension = "csv"
    else:
        try:
            payload = _build_xlsx_payload(
                diffs,
                columns=columns,
                summary=export_summary,
                sheet_mode=sheet_mode,
            )
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        media_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        extension = "xlsx"

    filename = _export_filename(req.filename, extension)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "X-BOM-Compare-Added": str(export_summary.get("added", 0)),
        "X-BOM-Compare-Removed": str(export_summary.get("removed", 0)),
        "X-BOM-Compare-Modified": str(export_summary.get("modified", 0)),
        "X-BOM-Compare-Unchanged": str(export_summary.get("unchanged", 0)),
    }
    return Response(
        content=payload,
        media_type=media_type,
        headers=headers,
    )


@router.post("/apply", response_model=BomApplyResponse)
def apply_bom_changes(
    req: BomApplyRequest,
    db: Session = Depends(_get_db),
    current_user: Any = Depends(_current_user),
) -> BomApplyResponse:
    from yuantus.meta_engine.models.item import Item
    from yuantus.meta_engine.schemas.aml import AMLAction, GenericItem
    from yuantus.meta_engine.services.bom_service import BOMService
    from yuantus.meta_engine.services.engine import AMLEngine

    identity, roles = _identity_from_user(current_user)
    engine = AMLEngine(db, identity_id=identity, roles=roles)
    bom_service = BOMService(db)

    results: List[Dict[str, Any]] = []
    dry_run = bool(req.dry_run)
    preview = bool(req.preview)
    if preview:
        for change in req.changes:
            op = change.op.strip().lower()
            if op in {"remove", "delete"}:
                if not change.relationship_id:
                    raise HTTPException(
                        status_code=400, detail="relationship_id is required for remove"
                    )
                rel_item = db.get(Item, change.relationship_id)
                if not rel_item:
                    raise HTTPException(status_code=404, detail="Relationship not found")
                results.append(
                    {
                        "op": "remove",
                        "relationship_id": rel_item.id,
                        "item_type": rel_item.item_type_id,
                    }
                )
                continue

            if op == "update":
                if not change.relationship_id:
                    raise HTTPException(
                        status_code=400, detail="relationship_id is required for update"
                    )
                rel_item = db.get(Item, change.relationship_id)
                if not rel_item:
                    raise HTTPException(status_code=404, detail="Relationship not found")
                props = dict(change.properties or {})
                if change.quantity is not None:
                    props["quantity"] = change.quantity
                current_props = dict(rel_item.properties or {})
                merged = dict(current_props)
                merged.update(props)
                results.append(
                    {
                        "op": "update",
                        "relationship_id": rel_item.id,
                        "item_type": rel_item.item_type_id,
                        "current_properties": current_props,
                        "next_properties": merged,
                    }
                )
                continue

            if op == "add":
                if not change.parent_id or not change.child_id:
                    raise HTTPException(
                        status_code=400,
                        detail="parent_id and child_id are required for add",
                    )
                if bom_service.detect_cycle(change.parent_id, change.child_id):
                    raise HTTPException(
                        status_code=409,
                        detail="Adding this relationship would create a cycle",
                    )
                parent_item = db.get(Item, change.parent_id)
                if not parent_item:
                    raise HTTPException(status_code=404, detail="Parent item not found")
                child_item = db.get(Item, change.child_id)

                rel_type = change.relationship_type
                if not rel_type:
                    rel_type = (req.relationship_types or ["Part BOM"])[0]

                props = dict(change.properties or {})
                props["related_id"] = change.child_id
                if change.quantity is not None:
                    props["quantity"] = change.quantity

                results.append(
                    {
                        "op": "add",
                        "relationship_type": rel_type,
                        "parent_id": parent_item.id,
                        "parent_item_type": parent_item.item_type_id,
                        "child_id": change.child_id,
                        "child_item_type": child_item.item_type_id
                        if child_item
                        else None,
                        "child_exists": bool(child_item),
                        "properties": props,
                    }
                )
                continue

            raise HTTPException(status_code=400, detail=f"Unknown op: {change.op}")

        return BomApplyResponse(ok=True, results=results, dry_run=True, preview=True)
    try:
        for change in req.changes:
            op = change.op.strip().lower()
            if op in {"remove", "delete"}:
                if not change.relationship_id:
                    raise HTTPException(
                        status_code=400, detail="relationship_id is required for remove"
                    )
                rel_item = db.get(Item, change.relationship_id)
                if not rel_item:
                    raise HTTPException(status_code=404, detail="Relationship not found")
                aml = GenericItem(
                    id=rel_item.id,
                    type=rel_item.item_type_id,
                    action=AMLAction.delete,
                )
                result = engine.apply(aml)
                results.append({"op": "remove", "result": result})
                continue

            if op == "update":
                if not change.relationship_id:
                    raise HTTPException(
                        status_code=400, detail="relationship_id is required for update"
                    )
                rel_item = db.get(Item, change.relationship_id)
                if not rel_item:
                    raise HTTPException(status_code=404, detail="Relationship not found")
                props = dict(change.properties or {})
                if change.quantity is not None:
                    props["quantity"] = change.quantity
                aml = GenericItem(
                    id=rel_item.id,
                    type=rel_item.item_type_id,
                    action=AMLAction.update,
                    properties=props,
                )
                result = engine.apply(aml)
                results.append({"op": "update", "result": result})
                continue

            if op == "add":
                if not change.parent_id or not change.child_id:
                    raise HTTPException(
                        status_code=400,
                        detail="parent_id and child_id are required for add",
                    )
                if bom_service.detect_cycle(change.parent_id, change.child_id):
                    raise HTTPException(
                        status_code=409,
                        detail="Adding this relationship would create a cycle",
                    )
                parent_item = db.get(Item, change.parent_id)
                if not parent_item:
                    raise HTTPException(status_code=404, detail="Parent item not found")

                rel_type = change.relationship_type
                if not rel_type:
                    rel_type = (req.relationship_types or ["Part BOM"])[0]

                props = dict(change.properties or {})
                props["related_id"] = change.child_id
                if change.quantity is not None:
                    props["quantity"] = change.quantity

                aml = GenericItem(
                    type=rel_type,
                    action=AMLAction.add,
                    properties=props,
                )
                result = engine.apply_relationship(aml, source_item=parent_item)
                results.append({"op": "add", "result": result})
                continue

            raise HTTPException(status_code=400, detail=f"Unknown op: {change.op}")

        if dry_run:
            db.rollback()
        else:
            db.commit()
    except PLMException as exc:
        db.rollback()
        raise HTTPException(status_code=exc.status_code, detail=exc.to_dict()) from exc
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return BomApplyResponse(ok=True, results=results, dry_run=dry_run, preview=False)
